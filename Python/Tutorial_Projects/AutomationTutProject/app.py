import openpyxl as xl

# This program changes all the prices in the price column of the
# transactions.xlsx sheet to 10% less it's current value.

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
# cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save("transactions2.xlsx")
